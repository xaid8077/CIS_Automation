"""
excel_writer.py
───────────────
Writes payload into the IL.xlsx and IO.xlsx templates.

─── Instrument List (IL.xlsx) ───────────────────────────────────────────────
Cover sheet     → fills project metadata into fixed cells.
                  AI14 receives the Instrument List document number
                  from fi_meta["docNumber"], not from the global header.

Instrument List → Section 1 data starting at row 6.
                  Column A   : serial numbers (1, 2, 3 …)
                  Column B   : generated instrument code from Instrument Name
                  Columns C–S: instrument data (fixed mapping per spec).
                  Column I   : intentionally skipped (spacer in template).
                  Columns U–V: velocity & optimised diameter for flowmeters.
                  Sections 2 & 3 reserved for a future update.

─── IO List (IO.xlsx) ───────────────────────────────────────────────────────
Cover sheet     → same cell mapping as IL cover.
                  AI14 receives the IO List document number from io_meta.

IO List sheet   → Three section blocks starting at row 6.
                  Each block opens with a merged header row (A:N).
                  Column A : serial number (continuous across all sections)
                  Column B : Tag No
                  Column C : Instrument Name
                  Column D : Service Description
                  Column E : Signal Type
                  Column F : Source
                  Column G : Destination
                  Column H : DI  — 1 if Signal == "DI", else blank
                  Column I : DO  — 1 if Signal == "DO", else blank
                  Column J : AI  — 1 if Signal == "AI", else blank
                  Column K : AO  — 1 if Signal == "AO", else blank
                  Column L : skipped (blank, bordered)
                  Column M : TRIP — "TRIP" if "trip" in Service Description
"""

import json
import math
import re
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side


# ─── Style helpers ────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border(left="thin", right="thin", top="thin", bottom="thin") -> Border:
    def s(st):
        return Side(style=st) if st else Side()
    return Border(left=s(left), right=s(right), top=s(top), bottom=s(bottom))


THIN        = _border()
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ─── Paths ────────────────────────────────────────────────────────────────────
# BASE_DIR is the project root (two levels up from utils/excel_writer.py).

BASE_DIR      = Path(__file__).resolve().parent.parent
TEMPLATE_PATH = BASE_DIR / "templates" / "IL.xlsx"
CODE_MAP_PATH = BASE_DIR / "templates" / "instrument_code_map.json"
IO_TEMPLATE_PATH = BASE_DIR / "templates" / "IO.xlsx"


# ─── Flowmeter keywords ───────────────────────────────────────────────────────
# Any instrument whose name contains one of these strings (case-insensitive)
# will have velocity and optimised diameter calculated and written to U & V.
# Add more keywords here as your naming conventions require.

FLOW_KEYWORDS = [
    "flow transmitter",
    "flowmeter",
    "flow meter",
    "flow element",
    "magnetic flowmeter",
    "magnetic flow meter",
    "electromagnetic flow",
    "vortex flow",
    "turbine flow",
    "ultrasonic flow",
    "coriolis",
    "rotameter",
]

# Standard nominal bore sizes in mm (DN series).
# The velocity optimiser steps through this list so every result lands on
# a real pipe size.  DN5000 is NOT included — it is not a recognised
# standard nominal bore in ASME B36.10, ISO 4200, or DIN EN 10220.
STANDARD_DN = [
    15, 20, 25, 32, 40, 50, 65, 80, 100, 125, 150,
    200, 250, 300, 350, 400, 450, 500, 600, 700, 800,
    900, 1000, 1200, 1400, 1600, 1800, 2000, 2500, 3000,
]

# Velocity band for liquid service (m/s).
V_MIN = 0.2
V_MAX = 5.0

# Cover sheet: starting column index (1-based) for character-by-character
# document number writing.  Column D = 4.
# *** Verify this against your template before deploying. ***
DOC_NUMBER_ROW       = 44
DOC_NUMBER_START_COL = 4   # D


# ─── Instrument code config — lazy loader ─────────────────────────────────────
# The config is NOT loaded at module import time.  Loading it on first use
# (inside write_workbook) means a missing JSON file will only raise an error
# during an actual download request — not at Flask startup — so the rest of
# the application continues to function normally.

_instrument_code_config: Optional[Dict[str, Any]] = None   # cache


def _normalize_text(text: str) -> str:
    """
    Normalise text for reliable exact/contains matching.
    Lowercases, replaces & with 'and', collapses punctuation and
    extra whitespace to single spaces.
    """
    text = str(text or "").strip().lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[-_/(),]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text


def _load_instrument_code_config(path: Path) -> Dict[str, Any]:
    """
    Read and validate instrument_code_map.json.

    Expected JSON structure:
    {
        "exact_map":      { "pressure transmitter": "PT", ... },
        "contains_rules": [ ["flow transmitter", "FT"], ... ],
        "stop_words":     ["and", "of", "for", ...]
    }

    Keys in exact_map and keywords in contains_rules are normalised at
    load time so lookup comparisons are always against the same form.

    Raises FileNotFoundError if the file is missing.
    Raises ValueError if the structure is invalid.
    """
    if not path.exists():
        raise FileNotFoundError(
            f"Instrument code map not found at {path}.\n"
            "Place instrument_code_map.json inside the templates/ folder."
        )

    with path.open("r", encoding="utf-8") as f:
        config = json.load(f)

    if not isinstance(config, dict):
        raise ValueError("instrument_code_map.json must be a JSON object.")

    config.setdefault("exact_map",      {})
    config.setdefault("contains_rules", [])
    config.setdefault("stop_words",     [])

    if not isinstance(config["exact_map"], dict):
        raise ValueError("'exact_map' must be a JSON object.")
    if not isinstance(config["contains_rules"], list):
        raise ValueError("'contains_rules' must be a JSON array.")
    if not isinstance(config["stop_words"], list):
        raise ValueError("'stop_words' must be a JSON array.")

    # Normalise exact_map keys once at load time.
    normalised_exact: Dict[str, str] = {
        _normalize_text(k): str(v)
        for k, v in config["exact_map"].items()
    }

    # Normalise contains_rules keywords once at load time.
    normalised_rules: List[Tuple[str, str]] = []
    for item in config["contains_rules"]:
        if not isinstance(item, list) or len(item) != 2:
            raise ValueError(
                "Each entry in 'contains_rules' must be a two-item array: "
                "[keyword, code]"
            )
        keyword, code = item
        normalised_rules.append((_normalize_text(keyword), str(code)))

    normalised_stops: set = {_normalize_text(w) for w in config["stop_words"]}

    return {
        "exact_map":      normalised_exact,
        "contains_rules": normalised_rules,
        "stop_words":     normalised_stops,
    }


def _get_code_config() -> Dict[str, Any]:
    """
    Return the instrument code config, loading it on first call (lazy load).
    Subsequent calls return the cached result without re-reading the file.
    """
    global _instrument_code_config
    if _instrument_code_config is None:
        _instrument_code_config = _load_instrument_code_config(CODE_MAP_PATH)
    return _instrument_code_config


# ─── Instrument code generation ───────────────────────────────────────────────

def _fallback_instrument_code(instr_name: str, stop_words: set) -> str:
    """
    Generate a best-effort acronym from meaningful words when no mapping
    exists in the config.  Takes the first letter of up to three
    non-stop words; numeric tokens are kept whole.

    This is only a fallback — results should be reviewed by an engineer.
    """
    words = [
        w for w in _normalize_text(instr_name).split()
        if w and w not in stop_words
    ]
    if not words:
        return ""

    parts = []
    for w in words[:3]:
        if re.search(r"\d", w):
            parts.append(w.upper())   # keep numeric tokens whole, e.g. "4-20ma"
        else:
            parts.append(w[0].upper())

    return "".join(parts)


def _generate_instrument_code(instr_name: str) -> str:
    """
    Derive the instrument type code from an instrument name.

    Priority order:
      1. Exact normalised match in exact_map.
      2. First matching contains_rule (order matters — put specific rules first).
      3. Fallback acronym from meaningful words.

    Returns an empty string for blank instrument names.
    """
    text = _normalize_text(instr_name)
    if not text:
        return ""

    config        = _get_code_config()
    exact_map     = config["exact_map"]
    contains_rules = config["contains_rules"]
    stop_words    = config["stop_words"]

    # 1. Exact match
    if text in exact_map:
        return exact_map[text]

    # 2. Contains rules (checked in order — put more specific rules first in JSON)
    for keyword, code in contains_rules:
        if keyword and keyword in text:
            return code

    # 3. Fallback acronym
    return _fallback_instrument_code(text, stop_words)


# ─── Cover sheet ──────────────────────────────────────────────────────────────

def _fill_cover(ws, header: Dict[str, str], meta: Dict[str, str]) -> None:
    """
    Write project metadata into fixed cells on a Cover sheet.

    Global project fields come from `header`.
    The document number comes from `meta["docNumber"]` and is written to AI14.
    It is also written character-by-character across row DOC_NUMBER_ROW
    starting at DOC_NUMBER_START_COL (D44), because the template uses
    individual cells per character in its cover layout.

    This function is shared by both the IL and IO workbook writers — pass
    fi_meta for the Instrument List, io_meta for the IO List.
    """
    mapping = {
        "AI6":  header.get("date",        ""),
        "AI7":  header.get("preparedBy",  ""),
        "AI8":  header.get("checkedBy",   ""),
        "AI9":  header.get("approvedBy",  ""),
        "AI10": header.get("revision",    ""),
        "AI11": header.get("projectName", ""),
        "AI12": header.get("client",      ""),
        "AI13": header.get("consultant",  ""),
        "AI14": meta.get("docNumber",     ""),
    }

    for cell_ref, value in mapping.items():
        ws[cell_ref] = value

    # Character-by-character write for the styled document number box.
    doc_number = str(meta.get("docNumber", "") or "")
    for i, ch in enumerate(doc_number):
        ws.cell(row=DOC_NUMBER_ROW, column=DOC_NUMBER_START_COL + i, value=ch)


# ─── Velocity optimiser ───────────────────────────────────────────────────────

def _velocity_at(dia_mm: float, flow_m3h: float) -> float:
    """
    Return fluid velocity (m/s) for a circular pipe.

    Formula:  v = Q / A
              Q = flow_m3h / 3600           [m³/s]
              A = π × (dia_mm / 1000)² / 4  [m²]
    """
    area = math.pi * ((dia_mm / 1000) ** 2) / 4
    return (flow_m3h / 3600) / area


def _calculate_optimized_velocity(
    initial_dia_mm: float,
    flow_m3h: float,
) -> Tuple[float, float]:
    """
    Walk the standard DN series to find the smallest bore whose velocity
    falls within [V_MIN, V_MAX] m/s.

    Strategy
    --------
    1. Snap initial_dia_mm to the nearest standard DN.
    2. If velocity is already in range, return immediately.
    3. Velocity > V_MAX (pipe too small) → step UP the DN list.
    4. Velocity < V_MIN (pipe too large) → step DOWN the DN list.
    5. If no standard size achieves the target, return the closest extreme
       so the engineer always sees a real number rather than a blank cell.

    Returns (velocity_m_per_s, recommended_dia_mm) both rounded to 3 d.p.
    """
    if flow_m3h <= 0 or initial_dia_mm <= 0:
        return 0.0, initial_dia_mm

    # Snap to nearest standard DN.
    current_dia = min(STANDARD_DN, key=lambda dn: abs(dn - initial_dia_mm))
    idx         = STANDARD_DN.index(current_dia)
    velocity    = _velocity_at(current_dia, flow_m3h)

    if V_MIN <= velocity <= V_MAX:
        return round(velocity, 3), float(current_dia)

    if velocity > V_MAX:
        # Too fast — step up to larger bores.
        for dn in STANDARD_DN[idx + 1:]:
            v = _velocity_at(dn, flow_m3h)
            if v <= V_MAX:
                return round(v, 3), float(dn)
        # Still too fast at largest standard size — return largest with its velocity.
        largest = STANDARD_DN[-1]
        return round(_velocity_at(largest, flow_m3h), 3), float(largest)

    # Too slow — step down to smaller bores.
    for dn in reversed(STANDARD_DN[:idx]):
        v = _velocity_at(dn, flow_m3h)
        if v >= V_MIN:
            return round(v, 3), float(dn)
    # Still too slow at smallest standard size — return smallest with its velocity.
    smallest = STANDARD_DN[0]
    return round(_velocity_at(smallest, flow_m3h), 3), float(smallest)


def _extract_first_number(text: str) -> float:
    """
    Extract the first numeric value (integer or decimal) from a string.

    Examples:
        "DN50"       → 50.0
        "12.5 m³/h"  → 12.5
        "2 inch"     → 2.0   ← caller is responsible for unit conversion

    Raises ValueError if no number is found in the string.
    """
    matches = re.findall(r"[-+]?\d*\.\d+|\d+", str(text))
    if not matches:
        raise ValueError(f"No numeric value found in {text!r}")
    return float(matches[0])


# ─── Instrument List sheet — Section 1 ───────────────────────────────────────

FI_COL_MAP = {
    "C": "Tag No",
    "D": "Instrument Name",
    "E": "Service Description",
    "F": "Line Size",
    "G": "Medium",
    "H": "Specification",
    # Column I intentionally skipped — spacer/merged region in template.
    "J": "Process Conn",
    "K": "Work Press",
    "L": "Work Flow",
    "M": "Work Level",
    "N": "Design Press",
    "O": "Design Flow",
    "P": "Design Level",
    "Q": "Set-point",
    "R": "Range",
    "S": "UOM",
}

FI_START_ROW = 6


def _write_fi_fixed(ws, rows: List[Dict[str, Any]]) -> None:
    """
    Write Section 1 (Field Instruments) rows into the Instrument List sheet.

    Per row:
      Column A   : serial number (1-based)
      Column B   : instrument code derived from Instrument Name
      Columns C–S: instrument data per FI_COL_MAP
      Columns U–V: (flowmeters only) calculated velocity and optimised DN

    Rows 1–5 (template header block) are never touched.
    """
    for i, row_dict in enumerate(rows):
        r      = FI_START_ROW + i
        serial = i + 1

        # ── Column A: Serial number ───────────────────────────────────────
        sn_cell           = ws[f"A{r}"]
        sn_cell.value     = serial
        sn_cell.alignment = CENTER_WRAP
        sn_cell.border    = THIN

        # ── Column B: Instrument code ─────────────────────────────────────
        instrument_name   = row_dict.get("Instrument Name", "")
        code_cell         = ws[f"B{r}"]
        code_cell.value   = _generate_instrument_code(instrument_name)
        code_cell.alignment = CENTER_WRAP
        code_cell.border  = THIN

        # ── Columns C–S: Instrument data ──────────────────────────────────
        for col_letter, field_key in FI_COL_MAP.items():
            cell           = ws[f"{col_letter}{r}"]
            cell.value     = row_dict.get(field_key, "")
            # Tag No (C) is centred; all other data columns are left-aligned.
            cell.alignment = CENTER_WRAP if col_letter == "C" else LEFT_WRAP
            cell.border    = THIN

        # ── Columns U–V: Flowmeter velocity calculation ───────────────────
        instr_lower  = str(instrument_name).lower()
        is_flowmeter = any(kw in instr_lower for kw in FLOW_KEYWORDS)

        if is_flowmeter:
            try:
                dia_mm   = _extract_first_number(row_dict.get("Line Size",   "0"))
                flow_m3h = _extract_first_number(row_dict.get("Design Flow", "0"))
                v_final, d_final = _calculate_optimized_velocity(dia_mm, flow_m3h)

                ws[f"U{r}"].value = v_final
                ws[f"V{r}"].value = d_final

                for col in ("U", "V"):
                    ws[f"{col}{r}"].alignment = CENTER_WRAP
                    ws[f"{col}{r}"].border    = THIN

            except (ValueError, ZeroDivisionError):
                # Input could not be parsed — write ERR so the engineer
                # knows a calculation was attempted but the Line Size or
                # Design Flow value could not be read numerically.
                for col in ("U", "V"):
                    cell           = ws[f"{col}{r}"]
                    cell.value     = "ERR"
                    cell.alignment = CENTER_WRAP
                    cell.border    = THIN


# ─── Instrument List orchestrator ─────────────────────────────────────────────

def _fill_instrument_list(ws, payload: Dict[str, Any]) -> None:
    """
    Entry point for writing the 'Instrument List' sheet.
    Sections 2 & 3 will be added in a future update.
    """
    _write_fi_fixed(ws, payload.get("field_instruments", []))


# ─── Sections 2 & 3 — reserved for future implementation ─────────────────────


# ─── IL Public entry point ────────────────────────────────────────────────────

def write_workbook(
    payload: Dict[str, Any],
    destination: Union[BytesIO, Path],
) -> None:
    """
    Load IL.xlsx template, fill Cover + Instrument List, save to destination.

    The instrument code config (instrument_code_map.json) is loaded here on
    first call via _get_code_config() — not at module import time — so a
    missing JSON file only raises an error when a download is actually
    attempted, leaving the rest of the application unaffected.

    Parameters
    ----------
    payload:
        Dictionary from app.py._build_payload(). Expected keys:
          "header"            – global project fields (projectName, client, …)
          "fi_meta"           – {"docName": …, "docNumber": …} for this workbook
          "field_instruments" – list of row dicts for Section 1
          "electrical"        – list of row dicts for Section 2 (future)
          "mov"               – list of row dicts for Section 3 (future)

    destination:
        BytesIO for in-memory streaming to the browser, or a Path to save
        directly to disk.

    Raises
    ------
    FileNotFoundError
        If IL.xlsx is not found at TEMPLATE_PATH, or
        if instrument_code_map.json is not found at CODE_MAP_PATH.
    ValueError
        If instrument_code_map.json is structurally invalid.
    """
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"Template not found at {TEMPLATE_PATH}.\n"
            "Place IL.xlsx inside the templates/ folder of the project."
        )

    # Trigger lazy load of the instrument code config here, before we touch
    # the workbook.  Any config error surfaces now with a clear message rather
    # than mid-write after the workbook has already been partially modified.
    _get_code_config()

    wb      = load_workbook(TEMPLATE_PATH)
    header  = payload.get("header",  {})
    fi_meta = payload.get("fi_meta", {})

    # ── Cover sheet ───────────────────────────────────────────────────────────
    if "Cover" in wb.sheetnames:
        _fill_cover(wb["Cover"], header, fi_meta)

    # ── Instrument List sheet ─────────────────────────────────────────────────
    if "Instrument List" in wb.sheetnames:
        _fill_instrument_list(wb["Instrument List"], payload)
    else:
        # Sheet missing from template — create it as a fallback.
        ws = wb.create_sheet("Instrument List")
        _fill_instrument_list(ws, payload)

    wb.save(destination)


# ═════════════════════════════════════════════════════════════════════════════
# IO List
# ═════════════════════════════════════════════════════════════════════════════

IO_START_ROW    = 6
_IO_BLOCK_FILL  = _fill("252a3a")   # surface-3 tint — visually distinct header

# Block header alignment: left-indented, not wrapped (single-line label)
_IO_BLOCK_ALIGN = Alignment(horizontal="left", vertical="center",
                             indent=1, wrap_text=False)


def _io_cell(ws, row: int, col: int, value: Any, alignment: Alignment) -> None:
    """Write value + border + alignment into a single IO List data cell."""
    c           = ws.cell(row=row, column=col)
    c.value     = value
    c.alignment = alignment
    c.border    = THIN


def _write_io_block_header(ws, row: int, label: str) -> None:
    """
    Write a merged section-label row spanning columns A (1) through N (14).

    openpyxl requires the border to be applied to every cell in the merged
    range individually — otherwise only the outer edge of the top-left cell
    is styled.
    """
    ws.merge_cells(f"A{row}:N{row}")

    hdr           = ws.cell(row=row, column=1)
    hdr.value     = label
    hdr.fill      = _IO_BLOCK_FILL
    hdr.alignment = _IO_BLOCK_ALIGN
    hdr.border    = THIN

    # Apply border to the remaining cells in the merged range so the full
    # outer border renders correctly across all columns.
    for col in range(2, 15):   # columns B(2) … N(14)
        ws.cell(row=row, column=col).border = THIN


def _write_io_list_sheet(ws, payload: Dict[str, Any]) -> None:
    """
    Write all three instrument sections into the 'IO List' sheet.

    Section order and block labels:
      1. Field Instruments
      2. Electrical Equipment
      3. Motor Operated Valves

    Each section opens with a merged header row.  Serial numbers run
    continuously from 1 across all three sections.

    Column layout (1-based):
      1  A  Serial number
      2  B  Tag No
      3  C  Instrument Name
      4  D  Service Description
      5  E  Signal Type
      6  F  Source
      7  G  Destination
      8  H  DI  — 1 if Signal == "DI", else blank
      9  I  DO  — 1 if Signal == "DO", else blank
     10  J  AI  — 1 if Signal == "AI", else blank
     11  K  AO  — 1 if Signal == "AO", else blank
     12  L  Skipped (blank, bordered)
     13  M  TRIP — "TRIP" if "trip" in Service Description, else blank
    """
    sections = [
        ("Field Instruments",     payload.get("field_instruments", [])),
        ("Electrical Equipment",  payload.get("electrical",        [])),
        ("Motor Operated Valves", payload.get("mov",               [])),
    ]

    row    = IO_START_ROW
    serial = 1

    for block_label, rows in sections:
        # ── Merged block header ───────────────────────────────────────────────
        _write_io_block_header(ws, row, block_label)
        row += 1

        # ── Data rows ─────────────────────────────────────────────────────────
        for rd in rows:
            tag     = rd.get("Tag No",              "")
            instr   = rd.get("Instrument Name",     "")
            service = rd.get("Service Description", "")
            sig_t   = rd.get("Signal Type",         "")
            source  = rd.get("Source",              "")
            dest    = rd.get("Destination",         "")
            signal  = rd.get("Signal",              "").strip().upper()

            _io_cell(ws, row,  1, serial,                          CENTER_WRAP)
            _io_cell(ws, row,  2, tag,                             CENTER_WRAP)
            _io_cell(ws, row,  3, instr,                           LEFT_WRAP)
            _io_cell(ws, row,  4, service,                         LEFT_WRAP)
            _io_cell(ws, row,  5, sig_t,                           CENTER_WRAP)
            _io_cell(ws, row,  6, source,                          CENTER_WRAP)
            _io_cell(ws, row,  7, dest,                            CENTER_WRAP)
            _io_cell(ws, row,  8, 1 if signal == "DI" else "",     CENTER_WRAP)
            _io_cell(ws, row,  9, 1 if signal == "DO" else "",     CENTER_WRAP)
            _io_cell(ws, row, 10, 1 if signal == "AI" else "",     CENTER_WRAP)
            _io_cell(ws, row, 11, 1 if signal == "AO" else "",     CENTER_WRAP)
            _io_cell(ws, row, 12, "",                              CENTER_WRAP)  # L skipped
            trip = "TRIP" if "trip" in service.lower() else ""
            _io_cell(ws, row, 13, trip,                            CENTER_WRAP)

            serial += 1
            row    += 1


# ─── IO Public entry point ────────────────────────────────────────────────────

def write_io_workbook(
    payload: Dict[str, Any],
    destination: Union[BytesIO, Path],
) -> None:
    """
    Load IO.xlsx template, fill Cover + IO List sheet, save to destination.

    Parameters
    ----------
    payload:
        Dictionary from app.py._build_payload(). Expected keys:
          "header"            – global project fields (projectName, client, …)
          "io_meta"           – {"docName": …, "docNumber": …} for the IO List
          "field_instruments" – list of row dicts (Section 1)
          "electrical"        – list of row dicts (Section 2)
          "mov"               – list of row dicts (Section 3)

    destination:
        BytesIO for in-memory streaming to the browser, or a Path to save
        directly to disk.

    Raises
    ------
    FileNotFoundError
        If IO.xlsx is not found at IO_TEMPLATE_PATH.
    """
    if not IO_TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"IO template not found at {IO_TEMPLATE_PATH}.\n"
            "Place IO.xlsx inside the templates/ folder of the project."
        )

    wb      = load_workbook(IO_TEMPLATE_PATH)
    header  = payload.get("header",  {})
    io_meta = payload.get("io_meta", {})

    # ── Cover sheet ───────────────────────────────────────────────────────────
    if "Cover" in wb.sheetnames:
        _fill_cover(wb["Cover"], header, io_meta)

    # ── IO List sheet ─────────────────────────────────────────────────────────
    if "IO List" in wb.sheetnames:
        _write_io_list_sheet(wb["IO List"], payload)
    else:
        ws = wb.create_sheet("IO List")
        _write_io_list_sheet(ws, payload)

    wb.save(destination)