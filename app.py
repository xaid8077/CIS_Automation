from datetime import datetime
from pathlib import Path
import traceback
from io import BytesIO
from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from utils.validator import validate_payload

app = Flask(__name__)

# --- Constants for Styling ---
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
CENTER_ALIGNED = Alignment(horizontal="center", vertical="center")
LEFT_ALIGNED_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# --- Helper Functions ---
def clean(value):
    return value.strip() if isinstance(value, str) else value

def get_list(form, key):
    return [clean(v) for v in form.getlist(key)]

def unpack_pfl(pfl_list, index):
    """Helper to unpack packed P/F/L strings (e.g., '10/50/20') safely."""
    vals = pfl_list[index].split('/') if index < len(pfl_list) else ["", "", ""]
    vals += [""] * (3 - len(vals)) # Ensure exactly 3 elements
    return vals

# --- Section Builders ---
def build_fi_section(form, prefix):
    """Section 1: Field Instruments (Process Data + Signal IO Data)"""
    tags = get_list(form, f"{prefix}TagNo[]")
    instruments = get_list(form, f"{prefix}Instrument[]")
    services = get_list(form, f"{prefix}ServiceDescription[]")
    line_sizes = get_list(form, f"{prefix}LineSize[]")
    mediums = get_list(form, f"{prefix}Medium[]")
    specs = get_list(form, f"{prefix}TypeSpec[]")
    conns = get_list(form, f"{prefix}ProcessConnection[]")
    working_vals = get_list(form, f"{prefix}WorkingValues[]")
    design_vals = get_list(form, f"{prefix}DesignValues[]")
    set_points = get_list(form, f"{prefix}SetPoint[]")
    ranges = get_list(form, f"{prefix}InstrumentRange[]")
    uoms = get_list(form, f"{prefix}Uom[]")
    
    # Signal IO fields
    sig_types = get_list(form, f"{prefix}SignalType[]")
    sources = get_list(form, f"{prefix}Source[]")
    destinations = get_list(form, f"{prefix}Destination[]")
    signals = get_list(form, f"{prefix}Signal[]")

    rows = []
    for i in range(len(tags)):
        w_vals = unpack_pfl(working_vals, i)
        d_vals = unpack_pfl(design_vals, i)

        row = {
            "Tag No": tags[i] if i < len(tags) else "",
            "Instrument Name": instruments[i] if i < len(instruments) else "",
            "Service Description": services[i] if i < len(services) else "",
            "Line Size": line_sizes[i] if i < len(line_sizes) else "",
            "Medium": mediums[i] if i < len(mediums) else "",
            "Work Press": w_vals[0], "Work Flow": w_vals[1], "Work Level": w_vals[2],
            "Design Press": d_vals[0], "Design Flow": d_vals[1], "Design Level": d_vals[2],
            "Specification": specs[i] if i < len(specs) else "",
            "Process Conn": conns[i] if i < len(conns) else "",
            "Set-point": set_points[i] if i < len(set_points) else "",
            "Range": ranges[i] if i < len(ranges) else "",
            "UOM": uoms[i] if i < len(uoms) else "",
            "Signal Type": sig_types[i] if i < len(sig_types) else "",
            "Source": sources[i] if i < len(sources) else "",
            "Destination": destinations[i] if i < len(destinations) else "",
            "Signal": signals[i] if i < len(signals) else ""
        }
        if any(row.values()):
            rows.append(row)
    return rows

def build_el_section(form, prefix):
    """Section 2: Electrical (Base Data + Signal IO Data ONLY)"""
    tags = get_list(form, f"{prefix}TagNo[]")
    instruments = get_list(form, f"{prefix}Instrument[]")
    services = get_list(form, f"{prefix}ServiceDescription[]")
    sig_types = get_list(form, f"{prefix}SignalType[]")
    sources = get_list(form, f"{prefix}Source[]")
    destinations = get_list(form, f"{prefix}Destination[]")
    signals = get_list(form, f"{prefix}Signal[]")

    rows = []
    for i in range(len(tags)):
        row = {
            "Tag No": tags[i] if i < len(tags) else "",
            "Instrument Name": instruments[i] if i < len(instruments) else "",
            "Service Description": services[i] if i < len(services) else "",
            "Signal Type": sig_types[i] if i < len(sig_types) else "",
            "Source": sources[i] if i < len(sources) else "",
            "Destination": destinations[i] if i < len(destinations) else "",
            "Signal": signals[i] if i < len(signals) else ""
        }
        if any(row.values()):
            rows.append(row)
    return rows

def build_mov_section(form, prefix):
    """Section 3: MOVs (Base Data + Signal IO Data ONLY)"""
    tags = get_list(form, f"{prefix}TagNo[]")
    instruments = get_list(form, f"{prefix}Instrument[]")
    services = get_list(form, f"{prefix}ServiceDescription[]")
    sig_types = get_list(form, f"{prefix}SignalType[]")
    sources = get_list(form, f"{prefix}Source[]")
    destinations = get_list(form, f"{prefix}Destination[]")
    signals = get_list(form, f"{prefix}Signal[]")

    rows = []
    for i in range(len(tags)):
        row = {
            "Tag No": tags[i] if i < len(tags) else "",
            "Instrument Name": instruments[i] if i < len(instruments) else "",
            "Service Description": services[i] if i < len(services) else "",
            "Signal Type": sig_types[i] if i < len(sig_types) else "",
            "Source": sources[i] if i < len(sources) else "",
            "Destination": destinations[i] if i < len(destinations) else "",
            "Signal": signals[i] if i < len(signals) else ""
        }
        if any(row.values()):
            rows.append(row)
    return rows

# --- Excel Sheet Writer ---
def write_sheet(ws, rows, title, project_info=None):
    """Refined sheet writer with auto-formatting and borders."""
    ws.title = title
    if not rows:
        ws.append(["No data available for this section"])
        return

    # 1. Add Project Header
    if project_info:
        ws.append(["Project:", project_info.get('projectName', '')])
        ws.append(["Client:", project_info.get('client', '')])
        ws.append(["Doc No:", project_info.get('documentNumber', '')])
        ws.append([]) # Spacer row

    # 2. Add Table Headers
    headers = list(rows[0].keys())
    ws.append(headers)
    
    header_row_idx = ws.max_row
    for cell in ws[header_row_idx]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # 3. Add Data Rows
    for row_dict in rows:
        row_values = [row_dict.get(h, "") for h in headers]
        ws.append(row_values)
        
        # Apply borders and alignment to each data cell
        for cell in ws[ws.max_row]:
            cell.border = THIN_BORDER
            if cell.column_letter == 'A': 
                cell.font = Font(bold=True)
                cell.alignment = CENTER_ALIGNED
            else:
                cell.alignment = LEFT_ALIGNED_WRAP

    # 4. Auto-Adjust Column Widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max(12, min(max_length + 2, 50))
        ws.column_dimensions[column].width = adjusted_width

    # 5. Freeze Panes
    ws.freeze_panes = f"A{header_row_idx + 1}"

# --- Application Routes ---
@app.route("/preview", methods=["POST"])
def preview():
    try:
        payload = {
            "header": {k: request.form.get(k, "") for k in ["projectName", "documentName", "client", "documentNumber"]},
            "field_instruments": build_fi_section(request.form, "fi"),
            "electrical": build_el_section(request.form, "el"),
            "mov": build_mov_section(request.form, "mov"),
        }
        
        errors = validate_payload(payload)
        
        if errors:
            return jsonify({"success": False, "errors": errors}), 400
            
        return jsonify({"success": True, "message": "Validation passed! No errors found."}), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/submit", methods=["POST"])
def submit():
    try:
        header_info = {k: request.form.get(k, "") for k in ["projectName", "documentName", "client", "documentNumber"]}
        payload = {
            "header": header_info,
            "field_instruments": build_fi_section(request.form, "fi"),
            "electrical": build_el_section(request.form, "el"),
            "mov": build_mov_section(request.form, "mov"),
        }
        
        output = BytesIO()
        wb = Workbook()
        
        # Write the sheets
        write_sheet(wb.active, payload["field_instruments"], "Field Instruments", header_info)
        write_sheet(wb.create_sheet("Electrical"), payload["electrical"], "Electrical", header_info)
        write_sheet(wb.create_sheet("MOV"), payload["mov"], "MOV", header_info)
        
        wb.save(output)
        output.seek(0)
        
        filename = f"Instrument_List_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename)
    
    except Exception as e:
        traceback.print_exc() 
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/")
def index(): 
    return render_template("index.html")

if __name__ == "__main__":
    app.run(debug=True)