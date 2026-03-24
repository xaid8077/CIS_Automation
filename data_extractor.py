import json
import csv
from copy import copy
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, GradientFill
from openpyxl.cell.cell import MergedCell


# ---------- helpers ----------

def safe_str(v):
    return "" if v is None else str(v)

def serialize_color(color):
    if color is None:
        return None
    return {
        "type": safe_str(getattr(color, "type", None)),
        "rgb": safe_str(getattr(color, "rgb", None)),
        "indexed": getattr(color, "indexed", None),
        "auto": getattr(color, "auto", None),
        "theme": getattr(color, "theme", None),
        "tint": getattr(color, "tint", None),
    }

def serialize_side(side):
    if side is None:
        return None
    return {
        "style": safe_str(getattr(side, "style", None)),
        "color": serialize_color(getattr(side, "color", None)),
    }

def serialize_border(border):
    if border is None:
        return None
    return {
        "left": serialize_side(border.left),
        "right": serialize_side(border.right),
        "top": serialize_side(border.top),
        "bottom": serialize_side(border.bottom),
        "diagonal": serialize_side(border.diagonal),
        "diagonalUp": getattr(border, "diagonalUp", None),
        "diagonalDown": getattr(border, "diagonalDown", None),
        "outline": getattr(border, "outline", None),
        "vertical": serialize_side(getattr(border, "vertical", None)),
        "horizontal": serialize_side(getattr(border, "horizontal", None)),
    }

def serialize_fill(fill):
    if fill is None:
        return None

    base = {
        "fill_type": safe_str(getattr(fill, "fill_type", None)),
    }

    if isinstance(fill, PatternFill):
        base.update({
            "fgColor": serialize_color(getattr(fill, "fgColor", None)),
            "bgColor": serialize_color(getattr(fill, "bgColor", None)),
            "patternType": safe_str(getattr(fill, "patternType", None)),
        })
    elif isinstance(fill, GradientFill):
        base.update({
            "type": safe_str(getattr(fill, "type", None)),
            "degree": getattr(fill, "degree", None),
            "left": getattr(fill, "left", None),
            "right": getattr(fill, "right", None),
            "top": getattr(fill, "top", None),
            "bottom": getattr(fill, "bottom", None),
            "stop": [str(s) for s in getattr(fill, "stop", [])] if getattr(fill, "stop", None) else None,
        })
    return base

def serialize_font(font):
    if font is None:
        return None
    return {
        "name": safe_str(font.name),
        "charset": getattr(font, "charset", None),
        "family": getattr(font, "family", None),
        "size": font.sz,
        "bold": font.b,
        "italic": font.i,
        "underline": safe_str(font.u),
        "strike": font.strike,
        "color": serialize_color(font.color),
        "outline": font.outline,
        "shadow": font.shadow,
        "condense": font.condense,
        "extend": font.extend,
        "superscript": getattr(font.vertAlign, "value", None) == "superscript" if font.vertAlign else False,
        "subscript": getattr(font.vertAlign, "value", None) == "subscript" if font.vertAlign else False,
        "vertAlign": safe_str(font.vertAlign),
        "scheme": safe_str(font.scheme),
    }

def serialize_alignment(alignment):
    if alignment is None:
        return None
    return {
        "horizontal": safe_str(alignment.horizontal),
        "vertical": safe_str(alignment.vertical),
        "text_rotation": alignment.textRotation,
        "wrap_text": alignment.wrapText,
        "shrink_to_fit": alignment.shrinkToFit,
        "indent": alignment.indent,
        "relative_indent": getattr(alignment, "relativeIndent", None),
        "justify_last_line": getattr(alignment, "justifyLastLine", None),
        "reading_order": getattr(alignment, "readingOrder", None),
        "merge_cell": getattr(alignment, "mergeCell", None),
    }

def serialize_protection(protection):
    if protection is None:
        return None
    return {
        "locked": protection.locked,
        "hidden": protection.hidden,
    }

def serialize_hyperlink(hyperlink):
    if hyperlink is None:
        return None
    return {
        "target": safe_str(getattr(hyperlink, "target", None)),
        "location": safe_str(getattr(hyperlink, "location", None)),
        "tooltip": safe_str(getattr(hyperlink, "tooltip", None)),
        "display": safe_str(getattr(hyperlink, "display", None)),
    }

def serialize_comment(comment):
    if comment is None:
        return None
    return {
        "author": safe_str(comment.author),
        "text": safe_str(comment.text),
        "width": getattr(comment, "width", None),
        "height": getattr(comment, "height", None),
    }

def get_merged_map(ws):
    merged_map = {}
    for merged_range in ws.merged_cells.ranges:
        for row in ws[merged_range.coord]:
            for cell in row:
                merged_map[cell.coordinate] = merged_range.coord
    return merged_map

def is_default_dimension(dim):
    """
    Returns True if there is likely no custom metadata worth exporting.
    """
    attrs = [
        getattr(dim, "hidden", None),
        getattr(dim, "outlineLevel", None),
        getattr(dim, "collapsed", None),
        getattr(dim, "style", None),
    ]
    return not any(v not in (None, False, 0, "") for v in attrs)

def style_present(cell):
    return (
        cell.has_style
        or cell.number_format != "General"
        or cell.comment is not None
        or cell.hyperlink is not None
    )


# ---------- extraction ----------

def extract_workbook_metadata(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=False)

    workbook_data = {
        "file": str(Path(xlsx_path).resolve()),
        "worksheets": [],
        "defined_names": [],
        "workbook_properties": {},
    }

    # workbook core-ish properties
    props = wb.properties
    workbook_data["workbook_properties"] = {
        "title": safe_str(props.title),
        "subject": safe_str(props.subject),
        "creator": safe_str(props.creator),
        "description": safe_str(props.description),
        "keywords": safe_str(props.keywords),
        "category": safe_str(props.category),
        "last_modified_by": safe_str(props.lastModifiedBy),
        "created": safe_str(props.created),
        "modified": safe_str(props.modified),
    }

    # defined names
    try:
        for dn in wb.defined_names.definedName:
            workbook_data["defined_names"].append({
                "name": safe_str(dn.name),
                "value": safe_str(dn.value),
                "localSheetId": getattr(dn, "localSheetId", None),
                "hidden": getattr(dn, "hidden", None),
            })
    except Exception:
        pass

    for ws in wb.worksheets:
        merged_map = get_merged_map(ws)

        sheet_data = {
            "sheet_name": ws.title,
            "sheet_state": ws.sheet_state,
            "tab_color": safe_str(getattr(ws.sheet_properties.tabColor, "rgb", None)) if ws.sheet_properties and ws.sheet_properties.tabColor else None,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "dimensions": ws.calculate_dimension(),
            "freeze_panes": safe_str(ws.freeze_panes),
            "auto_filter_ref": safe_str(ws.auto_filter.ref) if ws.auto_filter else None,
            "merged_ranges": [str(rng) for rng in ws.merged_cells.ranges],
            "print_area": safe_str(getattr(ws, "print_area", None)),
            "print_titles_rows": safe_str(getattr(ws, "print_title_rows", None)),
            "print_titles_cols": safe_str(getattr(ws, "print_title_cols", None)),
            "page_margins": {
                "left": ws.page_margins.left,
                "right": ws.page_margins.right,
                "top": ws.page_margins.top,
                "bottom": ws.page_margins.bottom,
                "header": ws.page_margins.header,
                "footer": ws.page_margins.footer,
            },
            "page_setup": {
                "orientation": safe_str(ws.page_setup.orientation),
                "paper_size": safe_str(ws.page_setup.paperSize),
                "scale": ws.page_setup.scale,
                "fit_to_width": ws.page_setup.fitToWidth,
                "fit_to_height": ws.page_setup.fitToHeight,
                "first_page_number": ws.page_setup.firstPageNumber,
                "use_first_page_number": ws.page_setup.useFirstPageNumber,
                "black_and_white": ws.page_setup.blackAndWhite,
                "draft": ws.page_setup.draft,
                "cell_comments": safe_str(ws.page_setup.cellComments),
                "errors": safe_str(ws.page_setup.errors),
                "horizontal_dpi": ws.page_setup.horizontalDpi,
                "vertical_dpi": ws.page_setup.verticalDpi,
                "copies": ws.page_setup.copies,
            },
            "sheet_view": {
                "show_grid_lines": getattr(ws.sheet_view, "showGridLines", None),
                "show_row_col_headers": getattr(ws.sheet_view, "showRowColHeaders", None),
                "zoom_scale": getattr(ws.sheet_view, "zoomScale", None),
                "zoom_scale_normal": getattr(ws.sheet_view, "zoomScaleNormal", None),
                "tab_selected": getattr(ws.sheet_view, "tabSelected", None),
                "right_to_left": getattr(ws.sheet_view, "rightToLeft", None),
            },
            "sheet_protection": {
                "sheet": ws.protection.sheet,
                "objects": ws.protection.objects,
                "scenarios": ws.protection.scenarios,
                "format_cells": ws.protection.formatCells,
                "format_columns": ws.protection.formatColumns,
                "format_rows": ws.protection.formatRows,
                "insert_columns": ws.protection.insertColumns,
                "insert_rows": ws.protection.insertRows,
                "insert_hyperlinks": ws.protection.insertHyperlinks,
                "delete_columns": ws.protection.deleteColumns,
                "delete_rows": ws.protection.deleteRows,
                "select_locked_cells": ws.protection.selectLockedCells,
                "sort": ws.protection.sort,
                "auto_filter": ws.protection.autoFilter,
                "pivot_tables": ws.protection.pivotTables,
                "select_unlocked_cells": ws.protection.selectUnlockedCells,
            },
            "header_footer": {
                "oddHeader": safe_str(ws.oddHeader.left.text if ws.oddHeader and ws.oddHeader.left else None) + " | "
                             + safe_str(ws.oddHeader.center.text if ws.oddHeader and ws.oddHeader.center else None) + " | "
                             + safe_str(ws.oddHeader.right.text if ws.oddHeader and ws.oddHeader.right else None),
                "oddFooter": safe_str(ws.oddFooter.left.text if ws.oddFooter and ws.oddFooter.left else None) + " | "
                             + safe_str(ws.oddFooter.center.text if ws.oddFooter and ws.oddFooter.center else None) + " | "
                             + safe_str(ws.oddFooter.right.text if ws.oddFooter and ws.oddFooter.right else None),
            },
            "row_dimensions": [],
            "column_dimensions": [],
            "cells": [],
        }

        # Row dimensions
        for idx, dim in ws.row_dimensions.items():
            if dim.height is not None or not is_default_dimension(dim):
                sheet_data["row_dimensions"].append({
                    "row": idx,
                    "height": dim.height,
                    "hidden": dim.hidden,
                    "outline_level": getattr(dim, "outlineLevel", None),
                    "collapsed": dim.collapsed,
                    "style_id": getattr(dim, "style", None),
                    "custom_height": getattr(dim, "customHeight", None),
                })

        # Column dimensions
        for key, dim in ws.column_dimensions.items():
            if dim.width is not None or not is_default_dimension(dim):
                sheet_data["column_dimensions"].append({
                    "column": key,
                    "min": dim.min,
                    "max": dim.max,
                    "width": dim.width,
                    "hidden": dim.hidden,
                    "best_fit": getattr(dim, "bestFit", None),
                    "outline_level": getattr(dim, "outlineLevel", None),
                    "collapsed": dim.collapsed,
                    "style_id": getattr(dim, "style", None),
                    "custom_width": getattr(dim, "customWidth", None),
                })

        # Cell metadata
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue

                has_value = cell.value is not None
                has_style = style_present(cell)

                if not has_value and not has_style:
                    continue

                row_dim = ws.row_dimensions[cell.row]
                col_dim = ws.column_dimensions[cell.column_letter]

                cell_record = {
                    "sheet": ws.title,
                    "cell": cell.coordinate,
                    "row": cell.row,
                    "column": cell.column,
                    "column_letter": cell.column_letter,
                    "data_type": cell.data_type,
                    "is_date": cell.is_date,
                    "value": cell.value,
                    "display_value_str": safe_str(cell.value),
                    "formula": cell.value if cell.data_type == "f" else None,
                    "has_style": cell.has_style,
                    "style_id": cell.style_id,
                    "number_format": cell.number_format,
                    "font": serialize_font(copy(cell.font)),
                    "fill": serialize_fill(copy(cell.fill)),
                    "border": serialize_border(copy(cell.border)),
                    "alignment": serialize_alignment(copy(cell.alignment)),
                    "protection": serialize_protection(copy(cell.protection)),
                    "hyperlink": serialize_hyperlink(cell.hyperlink),
                    "comment": serialize_comment(cell.comment),
                    "merged_range": merged_map.get(cell.coordinate),
                    "row_height": row_dim.height,
                    "column_width": col_dim.width,
                }

                sheet_data["cells"].append(cell_record)

        workbook_data["worksheets"].append(sheet_data)

    return workbook_data


# ---------- export ----------

def write_csvs(metadata, out_dir):
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # workbook summary
    with open(out_dir / "workbook_summary.json", "w", encoding="utf-8") as f:
        json.dump(metadata["workbook_properties"], f, indent=2, ensure_ascii=False, default=str)

    # defined names
    if metadata.get("defined_names"):
        with open(out_dir / "defined_names.csv", "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=["name", "value", "localSheetId", "hidden"]
            )
            writer.writeheader()
            writer.writerows(metadata["defined_names"])

    # per sheet outputs
    for ws in metadata["worksheets"]:
        safe_sheet = "".join(c if c.isalnum() or c in (" ", "_", "-") else "_" for c in ws["sheet_name"]).strip()
        sheet_dir = out_dir / safe_sheet
        sheet_dir.mkdir(exist_ok=True)

        # sheet settings
        settings = {k: v for k, v in ws.items() if k not in ("cells", "row_dimensions", "column_dimensions")}
        with open(sheet_dir / "sheet_settings.json", "w", encoding="utf-8") as f:
            json.dump(settings, f, indent=2, ensure_ascii=False, default=str)

        # row dimensions
        if ws["row_dimensions"]:
            with open(sheet_dir / "row_dimensions.csv", "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=["row", "height", "hidden", "outline_level", "collapsed", "style_id", "custom_height"]
                )
                writer.writeheader()
                writer.writerows(ws["row_dimensions"])

        # column dimensions
        if ws["column_dimensions"]:
            with open(sheet_dir / "column_dimensions.csv", "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=["column", "min", "max", "width", "hidden", "best_fit", "outline_level", "collapsed", "style_id", "custom_width"]
                )
                writer.writeheader()
                writer.writerows(ws["column_dimensions"])

        # flattened cells csv
        flattened_cells = []
        for c in ws["cells"]:
            flattened_cells.append({
                "sheet": c["sheet"],
                "cell": c["cell"],
                "row": c["row"],
                "column": c["column"],
                "column_letter": c["column_letter"],
                "data_type": c["data_type"],
                "is_date": c["is_date"],
                "value": c["display_value_str"],
                "formula": c["formula"],
                "style_id": c["style_id"],
                "number_format": c["number_format"],

                "font_name": c["font"]["name"] if c["font"] else None,
                "font_size": c["font"]["size"] if c["font"] else None,
                "bold": c["font"]["bold"] if c["font"] else None,
                "italic": c["font"]["italic"] if c["font"] else None,
                "underline": c["font"]["underline"] if c["font"] else None,
                "font_color_rgb": c["font"]["color"]["rgb"] if c["font"] and c["font"]["color"] else None,

                "fill_type": c["fill"]["fill_type"] if c["fill"] else None,
                "fill_fg_rgb": c["fill"]["fgColor"]["rgb"] if c["fill"] and c["fill"].get("fgColor") else None,

                "align_h": c["alignment"]["horizontal"] if c["alignment"] else None,
                "align_v": c["alignment"]["vertical"] if c["alignment"] else None,
                "wrap_text": c["alignment"]["wrap_text"] if c["alignment"] else None,
                "text_rotation": c["alignment"]["text_rotation"] if c["alignment"] else None,

                "border_left": c["border"]["left"]["style"] if c["border"] and c["border"]["left"] else None,
                "border_right": c["border"]["right"]["style"] if c["border"] and c["border"]["right"] else None,
                "border_top": c["border"]["top"]["style"] if c["border"] and c["border"]["top"] else None,
                "border_bottom": c["border"]["bottom"]["style"] if c["border"] and c["border"]["bottom"] else None,

                "locked": c["protection"]["locked"] if c["protection"] else None,
                "hidden_formula": c["protection"]["hidden"] if c["protection"] else None,

                "hyperlink_target": c["hyperlink"]["target"] if c["hyperlink"] else None,
                "comment_author": c["comment"]["author"] if c["comment"] else None,
                "comment_text": c["comment"]["text"] if c["comment"] else None,

                "merged_range": c["merged_range"],
                "row_height": c["row_height"],
                "column_width": c["column_width"],
            })

        if flattened_cells:
            with open(sheet_dir / "cells.csv", "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=list(flattened_cells[0].keys()))
                writer.writeheader()
                writer.writerows(flattened_cells)

        # full cell json
        with open(sheet_dir / "cells_full.json", "w", encoding="utf-8") as f:
            json.dump(ws["cells"], f, indent=2, ensure_ascii=False, default=str)


# ---------- main ----------

if __name__ == "__main__":
    INPUT_FILE = "D:\OneDrive - L&T Construction\A1. Cable Interconnection Schedules (CIS)\Cable Interconnection Schedule - Patiala\WTP\LE20M034-I-WS-CW-CI-5034 PATIALA CABLE INTERCONNECTION SCHEDULE FOR WTP.xlsm"           # change this
    OUTPUT_DIR = "excel_data"   # change this if needed

    metadata = extract_workbook_metadata(INPUT_FILE)

    # master JSON
    Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)
    with open(Path(OUTPUT_DIR) / "workbook_full_metadata.json", "w", encoding="utf-8") as f:
        json.dump(metadata, f, indent=2, ensure_ascii=False, default=str)

    # CSV/JSON folder export
    write_csvs(metadata, OUTPUT_DIR)

    print(f"Done. Metadata exported to: {OUTPUT_DIR}")